from openpyxl import load_workbook

excelfile = load_workbook('product.xlsx')
allshseets = excelfile.sheetnames
sheet = excelfile[allshseets[0]]

result = []

count = len(sheet['A']) #เลือก column ไหนก็ได้มานับว่ามีกี่แถว 

for i in range(2,count+1):
	a = sheet.cell(row=i, column=1).value
	b = sheet.cell(row=i, column=2).value
	c = sheet.cell(row=i, column=3).value
	data = [a,b,c]
	result.append(data)

############

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import wikipedia
import time

#ต้องโหลด selenium แล้วเอา chromedriver มาวางไว้ใน folder ของ python หากไท้ได้วางให้ใส่ path ในวงเว็บ chrome แทน
driver = webdriver.Chrome()

url_login = 'http://uncle-machine.com/login/'
driver.get(url_login)

username = driver.find_element_by_id('username')
username.send_keys('uuu@gmail.com')

password = driver.find_element_by_id('password')
password.send_keys('123')

button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
button.click()
time.sleep(2)

#fill data in product page 
url_addproduct = 'http://uncle-machine.com/addproduct/'
driver.get(url_addproduct)

for res in result:
	name = driver.find_element_by_id('name')
	name.send_keys(res[0])

	price = driver.find_element_by_id('price')
	price.send_keys(res[2])

	detail = driver.find_element_by_id('detail')
	detail.send_keys(res[1])

	button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
	button.click()
	time.sleep(1)
