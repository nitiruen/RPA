import wikipedia
import random

wikipedia.set_lang('th')
#print(wikipedia.summary('ส้ม'))

'''
a = wikipedia.page("kfc")
print(a.title)
print('----'*20)
print(a.url)
print('----'*20)
print(a.content)
'''

keyword = ['football', 'basketball', 'running']
allresult = []
allprice = []

for i in keyword:
	result = wikipedia.summary(i)
	#print(result)
	allresult.append(result)
	allprice.append(random.randint(100,1000))
	#print('-----'*20)

#print(allresult)
#print(allprice)

##################################

from openpyxl import Workbook

excelfile = Workbook()
sheet = excelfile.active

sheet['A1'].value = 'Keyword'
sheet['B1'].value = 'Detail'
sheet['C1'].value = 'Price'

for i,pd in enumerate(zip(keyword,allresult,allprice)):
	for j in range (len(keyword)):
		sheet.cell(row=i+2, column=j+1).value = pd[j]
	

'''
for i in range(2,50):
	sheet.cell(row=i,column=1).value = i
'''



excelfile.save('product.xlsx')